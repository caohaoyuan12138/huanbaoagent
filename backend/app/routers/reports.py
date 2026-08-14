"""
报告写作模块
"""
from fastapi import APIRouter, Depends, HTTPException, Body, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import io
import re

from app.db.database import get_db
from app.db.models import (
    ReportTemplate, ReportInstance, DeviceReading, PollutionLimit,
    EnterpriseStandard, Standard, ReportExportStat
)

router = APIRouter()


@router.get("/templates")
def list_templates(db: Session = Depends(get_db)):
    templates = db.query(ReportTemplate).all()
    return [
        {
            "id": t.id,
            "name": t.name,
            "type": t.type,
            "description": t.description,
            "fields": t.fields,
        }
        for t in templates
    ]


@router.post("/templates")
def create_template(template: dict, db: Session = Depends(get_db)):
    db_obj = ReportTemplate(**template)
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return db_obj


@router.post("/generate")
def generate_report(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    template_id = payload.get("template_id")
    params = payload.get("params", {})
    if not template_id:
        raise HTTPException(status_code=422, detail="template_id 不能为空")
    template = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    report_type = template.type

    if report_type == "daily_inspection":
        content = _generate_daily_inspection(template, params, db)
    elif report_type == "exceed_analysis":
        content = _generate_exceed_analysis(template, params, db)
    elif report_type == "compliance_check":
        content = _generate_compliance_check(template, params, db)
    elif report_type == "annual_report":
        content = _generate_annual_report(template, params, db)
    else:
        content = f"报告生成中（{report_type}），当前为模板版本。"

    instance = ReportInstance(
        template_id=template_id,
        params=params,
        content=content,
        status="generated",
        generated_at=datetime.now(),
    )
    db.add(instance)
    db.commit()
    db.refresh(instance)
    return {
        "id": instance.id,
        "template_id": template_id,
        "content": content,
        "status": instance.status,
        "generated_at": instance.generated_at,
    }


def _generate_daily_inspection(template: ReportTemplate, params: dict, db: Session) -> str:
    date_str = params.get("date", datetime.now().strftime("%Y-%m-%d"))
    workshop = params.get("workshop", "一号车间")
    inspector = params.get("inspector", "值班人员")

    equipment_list = params.get("equipment", [
        {"name": "RTO蓄热焚烧炉", "status": "正常"},
        {"name": "水洗塔", "status": "正常"},
        {"name": "活性炭吸附箱", "status": "正常"},
        {"name": "废水调节池", "status": "正常"},
        {"name": "COD在线监测仪", "status": "正常"},
    ])

    emission_data = params.get("emission_data", [])
    exceed_items = [e for e in emission_data if e.get("status") == "exceed"]

    content = f"""# 环保日常巡查报告

**报告编号**: {params.get("report_no", "RX-" + datetime.now().strftime("%Y%m%d"))}
**巡查日期**: {date_str}
**巡查车间**: {workshop}
**巡查人员**: {inspector}

---

## 一、巡查概况

本日对{workshop}进行例行环保巡查，主要内容包括废气治理设施运行状况、废水处理设施运行状况、污染物排放情况及环保台账记录。

## 二、治理设施运行情况

| 序号 | 设施名称 | 运行状态 | 备注 |
|------|---------|---------|------|
"""
    for i, eq in enumerate(equipment_list, 1):
        status_icon = "✅ 正常" if eq.get("status") == "正常" else "⚠️ 异常"
        content += f"| {i} | {eq['name']} | {status_icon} | {eq.get('remark', '')} |\n"

    content += """
## 三、污染物排放情况

"""
    if exceed_items:
        content += "### ⚠️ 异常排放记录\n\n"
        for item in exceed_items:
            content += f"- {item['factor']}: {item.get('value', '-')} {item.get('unit', '')}（超标）\n"
        content += "\n**建议**: 立即排查异常原因，采取整改措施。\n\n"
    else:
        content += "本日各项污染物排放数据正常，未发现超标情况。\n\n"

    content += """## 四、环保台账检查

- [x] 废气治理设施运行记录完整
- [x] 活性炭更换记录完整
- [x] 废水排放记录完整
- [ ] 危废转移记录需补充

## 五、整改建议

"""
    if exceed_items:
        content += "1. 针对异常排放点立即排查原因\n"
        content += "2. 加强对治理设施的巡检频次\n"
        content += "3. 如需停产检修，提前向环保部门报备\n"
    else:
        content += "1. 继续保持现有运维模式\n"
        content += "2. 按计划进行设备维护保养\n"
        content += "3. 关注下周天气变化对治理效果的影响\n"

    content += f"""
## 六、签字确认

巡查人员：____________    日期：{date_str}
负责人：____________    日期：{date_str}
"""
    return content


def _generate_exceed_analysis(template: ReportTemplate, params: dict, db: Session) -> str:
    date_str = params.get("date", datetime.now().strftime("%Y-%m-%d"))
    factor = params.get("factor", "COD")
    limit_value = params.get("limit_value", 50)
    actual_value = params.get("actual_value", 75)
    exceed_ratio = params.get("exceed_ratio", actual_value / limit_value) if limit_value > 0 else 1.0

    content = f"""# 污染物超标分析报告

**报告编号**: {params.get("report_no", "CB-" + datetime.now().strftime("%Y%m%d"))}
**分析日期**: {date_str}
**超标因子**: {factor}
**排放限值**: {limit_value} mg/L
**实测值**: {actual_value} mg/L
**超标倍数**: {round(exceed_ratio - 1, 2)} 倍

---

## 一、超标情况概述

{date_str}，{params.get("workshop", "一号车间")}排放口{factor}监测值为{actual_value}mg/L，超过GB 8978-1996规定的{limit_value}mg/L排放限值，超标{round(exceed_ratio - 1, 2)}倍。

## 二、超标原因分析

### 2.1 可能原因排查

"""
    possible_causes = params.get("possible_causes", [
        "废水处理设施运行参数异常",
        "生产工序排放浓度波动",
        "监测设备误差",
        "瞬时排放峰值",
    ])
    for i, cause in enumerate(possible_causes, 1):
        content += f"{i}. {cause}\n"

    content += """
### 2.2 根本原因判定

"""
    root_cause = params.get("root_cause", "生化处理单元进水负荷突然升高")
    content += f"经排查，本次超标的主要原因是：**{root_cause}**。\n\n"

    content += """## 三、影响评估

- 本次超标持续时长：约 {duration} 小时
- 排放总量估算：约 {total} kg
- 对周边环境影响：{impact}
- 是否存在累积风险：{accumulation}

## 四、整改措施

| 序号 | 整改措施 | 责任人 | 完成时限 |
|------|---------|--------|---------|
| 1 | {action1} | {person1} | {due1} |
| 2 | {action2} | {person2} | {due2} |
| 3 | 加强监测频次，每日增加2次手工监测 | {person3} | 立即执行 |

"""
    content = content.format(
        duration=params.get("duration", "2"),
        total=params.get("total", "15"),
        impact=params.get("impact", "暂未发现明显环境影响"),
        accumulation=params.get("accumulation", "否"),
        action1=params.get("action1", "调整生化池曝气量"),
        person1=params.get("person1", "运维负责人"),
        due1=params.get("due1", "24小时内"),
        action2=params.get("action2", "排查生产排水管路"),
        person2=params.get("person2", "生产主管"),
        due2=params.get("due2", "48小时内"),
        person3=params.get("person3", "环保员"),
    )

    content += """## 五、预防措施

1. 完善生产负荷预警机制，排放高浓度废水前需提前通知环保部门
2. 增加废水处理设施应急缓冲池容量
3. 定期校准监测设备，确保数据准确
4. 建立超标应急响应预案，明确处置流程

## 六、附件

- 监测数据记录表
- 治理设施运行记录
- 整改完成情况照片

---

编制人：____________    审核人：____________    批准人：____________
"""
    return content


def _generate_compliance_check(template: ReportTemplate, params: dict, db: Session) -> str:
    date_str = params.get("date", datetime.now().strftime("%Y-%m-%d"))
    factory_name = params.get("factory_name", "某某化工有限公司")
    industry = params.get("industry", "petrochemical")

    content = f"""# 环保合规排查报告

**报告编号**: {params.get("report_no", "HP-" + datetime.now().strftime("%Y%m%d"))}
**排查日期**: {date_str}
**企业名称**: {factory_name}
**所属行业**: {industry}

---

## 一、排查范围

本次排查覆盖企业以下方面：
- 废气排放合规性
- 废水排放合规性
- 固体废物管理
- 噪声排放
- 环保手续完整性
- 应急预案与演练

## 二、废气排放合规性

| 排放口 | 污染物 | 标准限值 | 最近监测值 | 达标情况 |
|--------|--------|---------|-----------|---------|
"""
    exhaust_items = params.get("exhaust_items", [
        {"name": "排气筒DA001", "factor": "VOCs", "limit": 60, "value": 45, "status": "达标"},
        {"name": "排气筒DA002", "factor": "SO₂", "limit": 100, "value": 30, "status": "达标"},
        {"name": "排气筒DA003", "factor": "NOx", "limit": 100, "value": 55, "status": "达标"},
    ])
    for item in exhaust_items:
        status_icon = "✅" if item["status"] == "达标" else "⚠️"
        content += f"| {item['name']} | {item['factor']} | {item['limit']} mg/m³ | {item['value']} mg/m³ | {status_icon} {item['status']} |\n"

    content += """
## 三、废水排放合规性

| 排放口 | 污染物 | 标准限值 | 最近监测值 | 达标情况 |
|--------|--------|---------|-----------|---------|
"""
    wastewater_items = params.get("wastewater_items", [
        {"name": "总排口", "factor": "COD", "limit": 50, "value": 38, "status": "达标"},
        {"name": "总排口", "factor": "NH₃-N", "limit": 5, "value": 3.2, "status": "达标"},
    ])
    for item in wastewater_items:
        status_icon = "✅" if item["status"] == "达标" else "⚠️"
        content += f"| {item['name']} | {item['factor']} | {item['limit']} mg/L | {item['value']} mg/L | {status_icon} {item['status']} |\n"

    content += """
## 四、存在问题

"""
    issues = params.get("issues", [
        {"level": "一般", "content": "部分危废仓库标识牌老化，建议更换", "deadline": "30日内整改"},
        {"level": "较重", "content": "RTO设备在线监测数据上传有延迟，需排查网络问题", "deadline": "7日内整改"},
    ])
    for issue in issues:
        level_color = "🟡" if issue["level"] == "一般" else "🟠"
        content += f"- {level_color} **{issue['level']}**：{issue['content']}（{issue['deadline']}）\n"

    content += """
## 五、整改建议

1. 对存在问题建立整改台账，明确责任人和完成时限
2. 加强日常环保巡检，做到问题早发现早处理
3. 定期开展环保法律法规培训，提升员工环保意识

## 六、结论

经排查，{factory_name}环保总体合规，{issue_count}项问题需整改。建议按计划完成整改后提交复查申请。

---

排查人员：____________    日期：{date_str}
"""
    content = content.format(factory_name=factory_name, issue_count=len(issues))
    return content


def _generate_annual_report(template: ReportTemplate, params: dict, db: Session) -> str:
    year = params.get("year", datetime.now().year)
    factory_name = params.get("factory_name", "某某化工有限公司")

    content = f"""# {year}年度环保报告

**企业名称**: {factory_name}
**报告周期**: {year}年1月1日 — {year}年12月31日

---

## 一、企业概况

{factory_name}主要生产化工产品，年排放量数据汇总如下：

## 二、污染物排放统计

### 2.1 废气排放

| 污染物 | 排放量(t) | 达标率(%) | 较上年变化 |
|--------|----------|----------|-----------|
| VOCs | {params.get("vocs_total", "12.5")} | {params.get("vocs_rate", "98.5")} | {params.get("vocs_change", "-3.2%")} |
| SO₂ | {params.get("so2_total", "5.8")} | {params.get("so2_rate", "99.1")} | {params.get("so2_change", "-1.5%")} |
| NOx | {params.get("nox_total", "18.3")} | {params.get("nox_rate", "97.8")} | {params.get("nox_change", "+0.5%")} |
| PM | {params.get("pm_total", "2.1")} | {params.get("pm_rate", "99.5")} | {params.get("pm_change", "-5.0%")} |

### 2.2 废水排放

| 污染物 | 排放量(t) | 达标率(%) | 较上年变化 |
|--------|----------|----------|-----------|
| COD | {params.get("cod_total", "8.6")} | {params.get("cod_rate", "99.2")} | {params.get("cod_change", "-4.1%")} |
| NH₃-N | {params.get("nh3n_total", "0.9")} | {params.get("nh3n_rate", "99.8")} | {params.get("nh3n_change", "-2.3%")} |

## 三、环保投资

| 项目 | 投资额(万元) |
|------|------------|
| 废气治理设施改造 | {params.get("investment_gas", "150")} |
| 废水处理设施升级 | {params.get("investment_water", "200")} |
| 在线监测系统 | {params.get("investment_monitor", "50")} |
| 合计 | {params.get("investment_total", "400")} |

## 四、存在不足与改进计划

"""
    shortcomings = params.get("shortcomings", [
        "部分老旧治理设施效率下降，需逐步更新",
        "环保信息化管理水平有待提升",
    ])
    for i, s in enumerate(shortcomings, 1):
        content += f"{i}. {s}\n"

    content += """
## 五、下年度工作计划

1. 完成VOCs治理设施提标改造
2. 建设智慧环保管理平台
3. 开展清洁生产审核
4. 加强环保人才队伍建设

---

编制部门：环保部    编制人：____________    批准人：____________
"""
    return content


@router.get("/instances")
def list_instances(
    template_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(ReportInstance)
    if template_id:
        query = query.filter(ReportInstance.template_id == template_id)
    if status:
        query = query.filter(ReportInstance.status == status)
    instances = query.order_by(ReportInstance.generated_at.desc()).all()
    return [
        {
            "id": i.id,
            "template_id": i.template_id,
            "params": i.params,
            "content": i.content,
            "status": i.status,
            "generated_at": i.generated_at,
        }
        for i in instances
    ]


@router.get("/instances/{instance_id}")
def get_instance(instance_id: int, db: Session = Depends(get_db)):
    instance = db.query(ReportInstance).filter(ReportInstance.id == instance_id).first()
    if not instance:
        raise HTTPException(status_code=404, detail="报告不存在")
    return {
        "id": instance.id,
        "template_id": instance.template_id,
        "params": instance.params,
        "content": instance.content,
        "status": instance.status,
        "generated_at": instance.generated_at,
    }


@router.delete("/instances/{instance_id}")
def delete_instance(instance_id: int, db: Session = Depends(get_db)):
    instance = db.query(ReportInstance).filter(ReportInstance.id == instance_id).first()
    if not instance:
        raise HTTPException(status_code=404, detail="报告不存在")
    db.delete(instance)
    db.commit()
    return {"message": "删除成功"}


@router.post("/seed")
def seed_templates(db: Session = Depends(get_db)):
    templates = [
        ReportTemplate(
            name="日常巡查报告",
            type="daily_inspection",
            description="环保日常巡查报告模板",
            fields=["date", "workshop", "inspector", "equipment", "emission_data", "report_no"],
        ),
        ReportTemplate(
            name="超标分析报告",
            type="exceed_analysis",
            description="污染物超标原因分析及整改措施报告",
            fields=["date", "factor", "limit_value", "actual_value", "exceed_ratio", "workshop", "report_no"],
        ),
        ReportTemplate(
            name="合规排查报告",
            type="compliance_check",
            description="环保合规性全面排查报告",
            fields=["date", "factory_name", "industry", "exhaust_items", "wastewater_items", "issues", "report_no"],
        ),
        ReportTemplate(
            name="年度环保报告",
            type="annual_report",
            description="年度环保工作总结报告",
            fields=["year", "factory_name"],
        ),
    ]
    for t in templates:
        existing = db.query(ReportTemplate).filter(ReportTemplate.name == t.name).first()
        if not existing:
            db.add(t)
    db.commit()
    return {"message": f"已初始化 {len(templates)} 个报告模板"}


# ── Export helpers ────────────────────────────────────────────────────────────

def _parse_markdown_lines(text: str) -> List[dict]:
    """Extract headings and paragraphs from markdown report content for Excel export."""
    rows = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("# "):
            rows.append({"level": 0, "text": stripped.lstrip("# ").strip()})
        elif stripped.startswith("## "):
            rows.append({"level": 1, "text": stripped.lstrip("# ").strip()})
        elif stripped.startswith("### "):
            rows.append({"level": 2, "text": stripped.lstrip("# ").strip()})
        else:
            rows.append({"level": 3, "text": stripped})
    return rows


def _generate_pdf(instance: ReportInstance, template: Optional[ReportTemplate]) -> bytes:
    """Generate a formatted PDF report using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, black
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    doc = SimpleDocTemplate(
        io.BytesIO(),
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=6 * cm,
        alignment=TA_CENTER,
        textColor=HexColor("#1a1a2e"),
        fontName="Helvetica-Bold",
    )
    heading2_style = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=12,
        spaceAfter=6,
        textColor=HexColor("#16213e"),
    )
    heading3_style = ParagraphStyle(
        "H3",
        parent=styles["Heading3"],
        fontSize=11,
        spaceBefore=8,
        spaceAfter=4,
        textColor=HexColor("#0f3460"),
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=6,
        leading=14,
        alignment=TA_JUSTIFY,
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontSize=9,
        spaceAfter=3,
        textColor=HexColor("#555555"),
    )

    elements = []

    # Title
    template_name = template.name if template else "环保报告"
    elements.append(Paragraph(template_name, title_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=HexColor("#0f3460"), spaceAfter=12))

    # Meta info
    meta_lines = [
        f"报告编号: {instance.params.get('report_no', f'REP-{instance.id}')} ",
        f"生成日期: {instance.generated_at.strftime('%Y-%m-%d %H:%M') if instance.generated_at else ''}",
        f"模板类型: {template.type if template else 'unknown'}",
    ]
    for line in meta_lines:
        if line.strip():
            elements.append(Paragraph(line, meta_style))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#cccccc"), spaceAfter=12))

    # Content processing
    content = instance.content or ""

    for line in content.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        # Skip table rows to handle them separately
        if stripped.startswith("|") and stripped.endswith("|"):
            continue

        if stripped.startswith("# "):
            elements.append(Paragraph(stripped.lstrip("# ").strip(), heading2_style))
        elif stripped.startswith("## "):
            elements.append(Paragraph(stripped.lstrip("# ").strip(), heading3_style))
        elif stripped.startswith("### "):
            elements.append(Paragraph(stripped.lstrip("# ").strip(), body_style))
        elif stripped.startswith("---"):
            elements.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#cccccc"), spaceAfter=6))
        elif stripped.startswith("- [x]") or stripped.startswith("- [X]"):
            elements.append(Paragraph(f"<b>✅</b> {stripped[7:].strip()}", body_style))
        elif stripped.startswith("- [ ]"):
            elements.append(Paragraph(f"<b>⬜</b> {stripped[7:].strip()}", body_style))
        elif stripped.startswith("- "):
            text = stripped[2:].strip()
            if text.startswith("**") and text.endswith("**"):
                text = f"<b>{text.strip('*')}</b>"
            elements.append(Paragraph(f"• {text}", body_style))
        elif re.match(r"^\d+\.", stripped):
            text = re.sub(r"^\d+\.\s*", "", stripped)
            if text.startswith("**") and text.endswith("**"):
                text = f"<b>{text.strip('*')}</b>"
            elements.append(Paragraph(f"{text}", body_style))
        elif stripped.startswith("|"):
            continue
        else:
            elements.append(Paragraph(stripped, body_style))

    # Process tables from markdown
    lines = content.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("|") and line.endswith("|"):
            # Collect table rows
            table_rows = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                row_text = lines[i].strip()
                cells = [c.strip() for c in row_text.split("|")[1:-1]]
                table_rows.append(cells)
                i += 1
            if len(table_rows) >= 2:
                header = table_rows[0]
                data = table_rows[1:]
                # Skip separator row if present
                if data and all(c.startswith("-") or c == "" for c in data[0]):
                    data = data[1:]
                if data:
                    t = Table([[Paragraph(c, body_style) for c in row] for row in data], repeatRows=1)
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0f3460")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cccccc")),
                        ("BACKGROUND", (0, 1), (-1, -1), HexColor("#f8f9fa")),
                    ]))
                    elements.append(Spacer(1, 0.2 * cm))
                    elements.append(t)
                    elements.append(Spacer(1, 0.3 * cm))
            continue
        i += 1

    elements.append(Spacer(1, 1 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#cccccc")))
    elements.append(Paragraph(
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  化工环保Agent v2.0",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8,
                       alignment=TA_CENTER, textColor=HexColor("#999999"))
    ))

    doc.build(elements)
    return doc.stream.getvalue()


def _generate_xlsx(instance: ReportInstance, template: Optional[ReportTemplate]) -> bytes:
    """Generate an Excel spreadsheet from report content using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "报告内容"

    header_font = Font(name="微软雅黑", bold=True, size=14, color="FFFFFF")
    section_font = Font(name="微软雅黑", bold=True, size=11, color="16213e")
    sub_font = Font(name="微软雅黑", bold=True, size=10, color="0f3460")
    body_font = Font(name="微软雅黑", size=10)
    meta_font = Font(name="微软雅黑", size=9, color="555555")
    title_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
    alt_fill = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )

    content = instance.content or ""
    lines = content.splitlines()
    row_idx = 1

    # Title row
    template_name = template.name if template else "环保报告"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value=template_name)
    title_cell.font = header_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    row_idx = 2

    # Meta info
    meta_data = [
        ("报告编号", instance.params.get("report_no", f"REP-{instance.id}")),
        ("生成日期", instance.generated_at.strftime("%Y-%m-%d %H:%M") if instance.generated_at else ""),
        ("模板类型", template.type if template else ""),
        ("报告状态", instance.status),
    ]
    for label, value in meta_data:
        if value:
            ws.cell(row=row_idx, column=1, value=label).font = meta_font
            ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="e8e8e8", end_color="e8e8e8", fill_type="solid")
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.cell(row=row_idx, column=3, value=value).font = body_font
            for c in range(1, 4):
                ws.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
    row_idx += 1

    # Parse content into structured rows
    table_buffer = []
    in_table = False

    def flush_table():
        nonlocal table_buffer, row_idx
        if len(table_buffer) >= 2:
            header = table_buffer[0]
            data = table_buffer[1:]
            # Skip separator rows
            if data and all(c.startswith("-") or c == "" for c in data[0]):
                data = data[1:]
            if data:
                for ci, h in enumerate(header, 1):
                    c = ws.cell(row=row_idx, column=ci, value=h)
                    c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center")
                    c.border = thin_border
                row_idx += 1
                for ri, row in enumerate(data):
                    for ci, val in enumerate(row, 1):
                        c = ws.cell(row=row_idx, column=ci, value=val)
                        c.font = body_font
                        c.border = thin_border
                        c.alignment = Alignment(horizontal="left")
                        if ri % 2 == 1:
                            c.fill = alt_fill
                    row_idx += 1
        table_buffer = []

    for line in lines:
        stripped = line.strip()
        if not stripped:
            if in_table:
                flush_table()
                in_table = False
            continue

        if stripped.startswith("|"):
            in_table = True
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            table_buffer.append(cells)
            continue
        else:
            if in_table:
                flush_table()
                in_table = False

        if stripped.startswith("# "):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value=stripped.lstrip("# ").strip())
            c.font = section_font
            c.fill = PatternFill(start_color="e8e8e8", end_color="e8e8e8", fill_type="solid")
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1
        elif stripped.startswith("## "):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value=stripped.lstrip("# ").strip())
            c.font = sub_font
            row_idx += 1
        elif stripped.startswith("### "):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value=stripped.lstrip("# ").strip())
            c.font = body_font
            c.fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
            row_idx += 1
        elif stripped == "---":
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            ws.cell(row=row_idx, column=1, value="").fill = PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")
            ws.row_dimensions[row_idx].height = 4
            row_idx += 1
        elif stripped.startswith("- [x]") or stripped.startswith("- [X]"):
            c = ws.cell(row=row_idx, column=1, value=f"✅ {stripped[7:].strip()}")
            c.font = body_font
            row_idx += 1
        elif stripped.startswith("- [ ]"):
            c = ws.cell(row=row_idx, column=1, value=f"⬜ {stripped[7:].strip()}")
            c.font = body_font
            row_idx += 1
        elif stripped.startswith("- "):
            c = ws.cell(row=row_idx, column=1, value=f"• {stripped[2:].strip()}")
            c.font = body_font
            row_idx += 1
        elif re.match(r"^\d+\.", stripped):
            text = re.sub(r"^\d+\.\s*", "", stripped)
            c = ws.cell(row=row_idx, column=1, value=text)
            c.font = body_font
            row_idx += 1
        else:
            c = ws.cell(row=row_idx, column=1, value=stripped)
            c.font = body_font
            row_idx += 1

    if in_table:
        flush_table()

    # Footer
    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    footer = ws.cell(row=row_idx, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  化工环保Agent v2.0")
    footer.font = Font(name="微软雅黑", size=8, color="999999")
    footer.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _track_export(db: Session, instance_id: int, fmt: str, size: int, success: bool):
    stat = ReportExportStat(
        report_instance_id=instance_id,
        format_type=fmt,
        file_size=size,
        success=success,
        created_at=datetime.now(),
    )
    db.add(stat)
    db.commit()


# ── Export endpoints ──────────────────────────────────────────────────────────

@router.post("/export/{report_id}")
def export_report(
    report_id: int,
    format: str = Query(..., regex="^(pdf|xlsx)$"),
    db: Session = Depends(get_db),
):
    instance = db.query(ReportInstance).filter(ReportInstance.id == report_id).first()
    if not instance:
        raise HTTPException(status_code=404, detail="报告不存在")
    if instance.status != "generated":
        raise HTTPException(status_code=400, detail=f"报告状态为 {instance.status}，无法导出")

    template = db.query(ReportTemplate).filter(ReportTemplate.id == instance.template_id).first()

    try:
        if format == "pdf":
            data = _generate_pdf(instance, template)
            filename = f"report_{report_id}.pdf"
            content_type = "application/pdf"
        else:
            data = _generate_xlsx(instance, template)
            filename = f"report_{report_id}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        _track_export(db, report_id, format, len(data), True)

        return Response(
            content=data,
            media_type=content_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        _track_export(db, report_id, format, 0, False)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/export/stats")
def export_stats(db: Session = Depends(get_db)):
    from sqlalchemy import func

    total_exports = db.query(ReportExportStat).count()
    pdf_count = db.query(ReportExportStat).filter(
        ReportExportStat.format_type == "pdf"
    ).count()
    xlsx_count = db.query(ReportExportStat).filter(
        ReportExportStat.format_type == "xlsx"
    ).count()
    success_count = db.query(ReportExportStat).filter(
        ReportExportStat.success == True
    ).count()
    fail_count = total_exports - success_count

    total_bytes = db.query(func.sum(ReportExportStat.file_size)).scalar() or 0

    recent = db.query(ReportExportStat).order_by(
        ReportExportStat.created_at.desc()
    ).limit(20).all()

    return {
        "total_exports": total_exports,
        "pdf_exports": pdf_count,
        "xlsx_exports": xlsx_count,
        "success_count": success_count,
        "fail_count": fail_count,
        "total_bytes": total_bytes,
        "recent_exports": [
            {
                "id": r.id,
                "report_instance_id": r.report_instance_id,
                "format": r.format_type,
                "file_size": r.file_size,
                "success": r.success,
                "created_at": r.created_at,
            }
            for r in recent
        ],
    }
